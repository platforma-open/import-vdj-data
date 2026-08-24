#!/usr/bin/env python3
"""Convert the first worksheet of a workbook to CSV.

A spreadsheet is a UI-boundary format: the platform's pipeline speaks csv/tsv and nothing
anywhere ingests a workbook as a data format. Converting here rather than in the browser means
no JavaScript xlsx dependency, and it means the header list, the identity-uniqueness check and
the import all read the same converted file rather than one of them re-parsing the original.

Kept deliberately close to immune-assay-data's converter, which is the shipped precedent.
"""

import argparse
import csv
import sys

from openpyxl import load_workbook


def find_header_row(rows: list) -> int:
    """Index of the first row that looks like a header.

    Workbooks routinely open with a title line or a blank one, and a naive read would make that
    the header and lose the real one. "More than one non-empty cell" is the same heuristic the
    precedent uses.
    """
    for i, row in enumerate(rows):
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
        if non_empty > 1:
            return i
    return 0


def xlsx_to_csv(input_file: str, output_file: str) -> None:
    wb = load_workbook(input_file, read_only=True, data_only=True)
    # First worksheet only. A workbook with the data on a later sheet converts to the wrong
    # thing silently, which is worth a picker eventually — the precedent carries the same TODO.
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)

    written = 0
    with open(output_file, "w", newline="") as f:
        writer = csv.writer(f)
        for row in rows[header_idx:]:
            writer.writerow(["" if cell is None else cell for cell in row])
            written += 1

    wb.close()
    print(f"Converted worksheet '{ws.title}' ({written} rows from row {header_idx + 1}) to {output_file}")


def main() -> None:
    p = argparse.ArgumentParser(description="Convert a workbook's first worksheet to CSV")
    p.add_argument("-i", "--input", required=True, help="Input workbook path")
    p.add_argument("-o", "--output", required=True, help="Output CSV path")
    args = p.parse_args()

    try:
        xlsx_to_csv(args.input, args.output)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
